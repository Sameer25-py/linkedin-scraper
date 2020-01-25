from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os 
from bs4 import BeautifulSoup as soup
import random
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import re

limit=1200
min=30
max=50

random.seed(random.random())

def domain_extracter(url):
	return url.split('@')[1]

def collect_soup(browser):
	button=browser.find_element_by_id('li-profile-name')
	button.click()
	browser.switch_to.window(browser.window_handles[1])
	time.sleep(2)
	link=browser.current_url
	browser.close()
	browser.switch_to.window(browser.window_handles[0])
	browser.get(link)
	from bs4 import BeautifulSoup as soup
	souper=soup(browser.page_source,'html.parser')
	return souper,link

def scraper_0(soup,browser,counter):
	name=''
	position=''
	company=''
	link=''
	domain=''
	email=''
	location=''
	name_finder1=soup.find('span',{'id':'li-profile-name'})
	if name_finder1!=None:
		#name =name_finder.string
		souper,link=collect_soup(browser)
		
		name_finder=souper.find('span',{'class':'profile-topcard-person-entity__name'})
		if name_finder is not None:
			name=name_finder.text.strip()
		
		position_finder=souper.find('span',{'class':'profile-topcard__summary-position-title'})
		if position_finder is not None:
			position=position_finder.text.strip()
    	
		company_finder=souper.find('div',{'class':'profile-topcard__summary-position'})
		if company_finder is not None:
			company_finder=company_finder.find('span',{'class':'Sans-14px-black-75%-bold'})
			if company_finder is not None:
				company=company_finder.text.strip()
		
		location_finder=souper.find('div',{'class':'profile-topcard__location-data'})
		if location_finder is not None:
			location=location_finder.text.strip()
		
		counter+=1

	else:
		print('not found ')
		return '','','','','',counter
	'''position_finder=souper.find('div',{'class':'li-user-title'})
	if position_finder!=None:	
		position=position_finder.next_element.split(' at')[0]
		
	company_finder=souper.find('span',{'class':'li-user-title-company'})
	if company_finder!=None:
		company=company_finder.string
	'''	
	return name,position,company,link,location,counter	
			
	
	#class=li-profile-name	
	
def scraper_1(souper,browser,counter):
	location=None
	name=None
	position=None
	company=None
	domain=''
	email=''
	domains=[]
	emails=[]
	lines=[]
	domain_re=r'(http[s]?://|www.)+([a-z.A-Z0-9]+$)'
	email_re=r"[\w\.-]+@[\w\.-]+"
	
	location_finder=souper.find('ul',{'class':'pv-top-card-v3--list-bullet'})
	if location_finder is not None:
		location=location_finder.find('li').text.strip()
	
	company_finder = souper.find('a',{'class':'pv-top-card-v3--experience-list-item'})
	if company_finder is not None:
		span= company_finder.find('span')
		if span is not None:
			company=span.text.strip()
            

	name_finder=souper.find('ul',{'class':'pv-top-card-v3--list'})
	if name_finder is not None:
		li=name_finder.find('li')
		if li is not None:
			name=li.text.strip()
			counter+=1

	about = souper.find('p')
	if about is not None:
		about=about.findAll('span')
	if about is not None:
		for tag in about:
			line=tag.text.strip()
			lines.append(line)
	x=souper.find('section',{'class':'pv-profile-section__card-item-v2'})
	if x is not None:
		info=x.find('p',{'class':'pv-entity__description'})
		if info is not None:
			info=info.text.strip()
			lines.append(info)
	for line in lines:
		link1=re.findall(domain_re,line)
		if link1:
			temp1=''.join(link1[0])
			domains.append(temp1)
		link2=re.findall(email_re,line)
		if link2:
			temp1=''.join(link2[0])
			emails.append(temp1)
	
	domain=','.join(domains)
	email=','.join(emails)

	position_finder =souper.find('section',{'class':'pv-profile-section__card-item-v2'})
	if position_finder is not None:
		pos2=position_finder.find('div',{'class':'pv-entity__summary-info-v2'})
		if pos2 is None:
			pos3=position_finder.find('div',{'class':'pv-entity__summary-info--background-section'})
			position=pos3.find('h3').text.strip()
		else:
			position=position_finder.find('div',{'class':'pv-entity__summary-info-v2'}).find('span').next_sibling.next_element.text.strip()                
	if domain=='': 
		from bs4 import BeautifulSoup as bs
		contact_addr=url + 'detail/contact-info/'
		browser.get(contact_addr)
		souper=bs(browser.page_source,'html.parser')
		souper=bs(browser.page_source,'html.parser')
		domain=souper.find('section',{'class':'ci-websites'})
		if domain is not None: 
			websites=domain.findAll('a')
			for website in websites:
				link1=re.findall(domain_re,website['href'])
				if link1:
					temp1=''.join(link1[0])
					domains.append(temp1)
	domain=','.join(domains)
	
	return name,position,company,location,domain,email,counter
		   
path=os.getcwd()
options= Options()
scraper_number=0
domain = None
name=None
position=None
comapny=None
link=None
wb=load_workbook(path + '\details.xlsx')
sheet=wb.active
options.add_experimental_option('debuggerAddress','127.0.0.1:9000')
counter = 0

browser=webdriver.Chrome(path + '\chromedriver',options=options)
browser.implicitly_wait(100)

for i in range(2,sheet.max_row + 1):
	print('----- PRESS CTRL + C TO EXIT -----')
	print('DAILY EMAIL SEARCHES LEFT ',limit-counter)
	url = sheet.cell(i,1).value
	if url is None:
		os.system('cls')
		print("ALL DONE")
		break
	if counter == limit:
		os.system('cls')
		#print("ALL DONE")
		break

	print('CURRENT URL {}'.format(url))
	if '@' in url:
		scraper_number=1
	else:
		scraper_number=2
	browser.get(url)
	souper = soup(browser.page_source,'html.parser')
	#print(scraper_number)
	if scraper_number == 1:
		domain=domain_extracter(url)
		name,position,company,link,location,counter=scraper_0(souper,browser,counter)
		#print('{},{},{},{}'.format(name,company,position,domain))
		sheet.cell(i,2).value=str(name)
		sheet.cell(i,3).value=str(company)
		sheet.cell(i,4).value=str(position)
		sheet.cell(i,5).value=str(domain)
		sheet.cell(i,6).value=str(location)
		sheet.cell(i,7).value=str(link)
		counter = counter
		wb.save(path + '\details.xlsx')
		print('done')
		wait=random.randint(min,max)
		print("WAITING FOR {} seconds".format(wait))
		time.sleep(wait)
		os.system('cls')
		#print('----- PRESS CTRL + C TO EXIT -----')

	elif scraper_number == 2:
		name,position,company,location,domain,email,counter=scraper_1(souper,browser,counter)
		#print('{},{},{},{}'.format(name,company,position,domain))
		sheet.cell(i,2).value=str(name)
		sheet.cell(i,3).value=str(company)
		sheet.cell(i,4).value=str(position)
		sheet.cell(i,5).value=str(domain)
		sheet.cell(i,6).value=str(location)
		sheet.cell(i,8).value=str(email)
		counter = counter 
		wb.save(path + '\details.xlsx')
		print('done')
		wait=random.randint(min,max)
		print("WAITING FOR {} seconds".format(wait))
		time.sleep(wait)
		os.system('cls')
		#print('----- PRESS CTRL + C TO EXIT -----')	

wb.save(path +'\details.xlsx')
wb.close()
print('ALL DONE !!')



	











